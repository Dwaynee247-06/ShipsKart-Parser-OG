"""
Utility script to generate a sample Excel workbook for testing.
Run from the project root:
  python scripts/create_sample_excel.py
"""
from pathlib import Path
import openpyxl

output = Path("sample_input.xlsx")
wb = openpyxl.Workbook()

# Sheet 1: Product requirements
ws1 = wb.active
ws1.title = "Sheet1"
ws1.append(["Sr No", "Product name", "Unit of measurement", "Quantity", "Remarks"])
ws1.append([1, "Cement", "KG", 100, None])
ws1.append([2, "Sand", "LTR", 200, "Course sand"])
ws1.append([3, "Gravel", "POUND", 150, None])
ws1.append([4, "Water", "LTR", 50, None])

# Sheet 2: Another table with same headers (will be grouped into requirements1)
ws2 = wb.create_sheet("Sheet2")
ws2.append(["Sr No", "Product name", "Unit of measurement", "Quantity", "Remarks"])
ws2.append([5, "Steel Rods", "KG", 500, "12mm dia"])
ws2.append([6, "Bricks", "Units", 1000, None])

# Sheet 3: Different headers (will become requirements2)
ws3 = wb.create_sheet("Sheet3")
ws3.append(["ID", "Vendor Name", "Contact", "City"])
ws3.append([1, "ABC Supplies", "9876543210", "Mumbai"])
ws3.append([2, "XYZ Traders", "9123456789", None])

wb.save(output)
print(f"Sample workbook saved to: {output}")
