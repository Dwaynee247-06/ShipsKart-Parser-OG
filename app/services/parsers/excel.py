"""
excel.py
--------
Parser for Excel workbooks (.xlsx, .xlsm, .xltx, .xltm).
All normalization, scoring, and table-building logic lives in _shared.py.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.parsers._shared import build_table_from_rows, clean_cell, normalize_header


def parse_excel(file_bytes: bytes) -> dict:
    """
    Parse every sheet in an Excel workbook.
    Returns a dict keyed by sheet name, each value being::

        {
            "document_info": { ...metadata above the table header... },
            "headers":       [ ...canonical column names... ],
            "rows":          [ { col: value, ... }, ... ],
        }
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    result: dict = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Convert the worksheet into a plain list-of-lists for _shared helpers
        raw_rows = [
            list(row)
            for row in ws.iter_rows(values_only=True)
        ]

        table = build_table_from_rows(raw_rows)
        if table is not None:
            result[sheet_name] = table

    return result
